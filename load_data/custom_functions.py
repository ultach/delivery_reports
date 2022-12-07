from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
import logging

def read_excel_table(
    file_path: str, 
    sheet_name: str, 
    is_smart_table: bool = True, 
    table_name: str = '', 
    promote_headers: bool = True, 
    rename: Dict[str, str] = {}, 
    convert: Dict[str, str] = {}, 
    metadata: Dict[str, Any] = {}
) -> Any: 

    open_file = load_workbook(file_path, data_only=True) 
    
    if sheet_name in open_file.sheetnames:
        sheet = open_file[sheet_name]
        
        ws_tables = [table[0] for table in open_file[sheet_name].tables.items()]
    else:
        raise KeyError(f'Отсутствует лист {sheet_name} в файле {file_path}')

    
    if is_smart_table:
        if table_name in ws_tables:
            table = sheet.tables[table_name] 
            table_range = table.ref
            source_ref = sheet[table_range]
        else:
            raise KeyError(f'Отсутствует таблица {table_name} на листе {sheet_name} в файле {file_path}')   
    else:
        source_ref = list(sheet.rows)

    if promote_headers:
        table_head = source_ref[0] 
        table_data = source_ref[1:] 
    else:
        first_row = source_ref[0]
        table_head = [f'column{str(i)}' for i in range(len(first_row))]
        table_data = source_ref
 
    new_headers = list(rename.keys())
    old_headers = [column.value for column in table_head]
 
    if new_headers == []:
        columns = old_headers
    else:
        columns = [rename.get(old_header, old_header) for old_header in old_headers]
    
    return_value = []
    for row in table_data: 
        row_val = []

        for index, cell in enumerate(row):
            column_name = columns[index]
            column_type = convert.get(column_name, None)
            if column_type == 'int':
                row_val.append(int(cell.value) if not cell.value is None else None)
            elif column_type == 'float':
                row_val.append(float(cell.value) if not cell.value is None else None)            
            else:
                row_val.append(cell.value)
        
        row_dict = { **metadata, **dict(zip(columns, row_val)) }
        return_value.append(row_dict)         
 

    return return_value


def get_emails(file_path):
    
    email_list = read_excel_table(file_path=file_path, sheet_name='Справочники', table_name='Email')
    
    email_dict = {
        'Success': {
            'To': [],
            'Cc': []
        },
        'All': {
            'To': [],
            'Cc': []
        }
    } 
    
    for email in email_list:
        email_address = email.get('Email адрес', '')
        email_type = email.get('Тип', '')
        recepient_type = email.get('Вид письма', '')
        
        if email_address is None: continue
        
        try:
            if recepient_type == 'Success':
                email_dict['Success'][email_type].append(email_address)
            elif recepient_type == 'All':
                email_dict['Success'][email_type].append(email_address)
                email_dict['All'][email_type].append(email_address)
            else:
                raise KeyError
        except KeyError as exc:
            logging.warning(f'E-mail адрес {email_address} не попал в рассылку, т.к. были нарушены правила заполнения справочника')
        
    logging.info('Списки email для отправки успешно получены')
    
    return email_dict